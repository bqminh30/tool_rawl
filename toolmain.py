import openpyxl
import time
from seleniumwire import webdriver
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import seleniumwire.undetected_chromedriver as uc
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
import random
import requests as req
from datetime import datetime

# Đường dẫn đến file Excel
excel_file_path = './Book2.xlsx'

# Mở file Excel
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active
max_row = sheet.max_row
exceptions_list = []
count = 1
# profile = webdriver.FirefoxProfile()
options = webdriver.FirefoxOptions()
# options.add_argument('ignore-certificate-errors')

for row in sheet.iter_rows(min_row=1, values_only=True):  # Bắt đầu từ hàng thứ 2 (hàng đầu tiên chứa tiêu đề)
    username, email, password, phone, code = row[:5]
    # the list of proxy to rotate on 
    PROXIES = [
        'http://113.161.131.43:80',
        'http://118.69.134.3:80',
        'http://118.69.134.0:80',
        'http://42.112.22.6:80',
        'http://103.74.121.88:3128',
        'http://14.177.235.17:8080',
        'http://116.111.119.16:8080',
        'http://14.241.62.12:19132',
        'http://113.161.93.29:8080',
        'http://117.4.50.142:32650',
    ]

    # # randomly extract a proxy
    random_proxy = random.choice(PROXIES)

    webdriver.DesiredCapabilities.FIREFOX['proxy'] = {
        "httpProxy":random_proxy,
        "ftpProxy":random_proxy,
        "sslProxy":random_proxy,
        "noProxy":None,
        "proxyType":"MANUAL",
    }
    options.add_argument('--proxy-server=%s' % random_proxy)
    now = datetime.now()
 
    print("now =", now)

# dd/mm/YY H:M:S
    dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
    print(count,' - ', random_proxy, ' - ', dt_string)
    driver = webdriver.Firefox(
        options=options,
    )
    # Điều hướng đến trang đăng ký
    driver.get('https://metahome.digital/sign-up')
    count = count+1
    time.sleep(1)
    
    # Tìm các phần tử input và nhập thông tin tương ứng
    driver.find_element(By.CSS_SELECTOR, '#signUp .box form input[name="email"]').send_keys(email)
    driver.find_element(By.CSS_SELECTOR, '#signUp .box form input[name="password"]').send_keys(password)
    driver.find_element(By.CSS_SELECTOR, '#signUp .box form input[name="confirmPassword"]').send_keys(password)
    driver.find_element(By.CSS_SELECTOR, '#signUp .box form input[name="referralCode"]').send_keys(code)
    element_to_click  = driver.find_element(By.CSS_SELECTOR, '#signUp .box form input[name="referralCode"]')
    # Điều hướng đến trang đăng ký 
    driver.execute_script("arguments[0].scrollIntoView();", element_to_click)
    time.sleep(1)
    # # Tìm và kiểm tra checkbox bằng ID
    wait = WebDriverWait(driver, 10)
    element = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#fullAgreement")))
    element.click()
        
    #Thực hiện đăng ký
    driver.find_element(By.CSS_SELECTOR, "#signUp .box form button.btn_signup").click()
    time.sleep(2)

    try:
        element_to_click  = driver.find_element(By.CSS_SELECTOR, '.btn_wait_kr')
        # Điều hướng đến trang đăng nhập 
        driver.execute_script("arguments[0].scrollIntoView();", element_to_click)
        # Nhập email và mật khẩu để đăng nhập
        driver.find_element(By.CSS_SELECTOR, '#logIn form input[name="email"]').send_keys(email)
        driver.find_element(By.CSS_SELECTOR, '#logIn form input[name="password"]').send_keys(password)
        # # Thực hiện đăng nhập
        driver.find_element(By.CSS_SELECTOR, "#logIn .box form button[type=submit]").click()

        time.sleep(1)
        driver.get('https://metahome.digital/mypage')
        wait = WebDriverWait(driver, 10)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.btn_add')))

        try:
            li_element = driver.find_element(By.XPATH, '//*[@id="myPage"]/section[2]/div/div/div/ul/li[2]/div/button').click()
            time.sleep(1)
            popup_element = driver.find_element(By.CSS_SELECTOR, '.register_phone form input[name="name"]').send_keys(username)
            popup_element = driver.find_element(By.CSS_SELECTOR, '.register_phone form input[name="phoneNumber"]').send_keys(phone)
            
            driver.find_element(By.CSS_SELECTOR, ".register_phone form button[type='submit']").click()

            time.sleep(1)
            driver.close()
            continue
            # 
        except:
            print(count)
    except:
        exceptions_list.append(str(count))
        with open("except1.txt", "a") as file:
            file.write(str(exceptions_list) + "\n")  # Ghi lỗi vào tệp văn bản
    # Kết thúc
    driver.quit()

